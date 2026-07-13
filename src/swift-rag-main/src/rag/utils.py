import ast
import json
from typing import Any, Callable, Dict, List

import numpy as np
import openpyxl
import pandas as pd
from llama_index.core import Document
from llama_index.core.schema import IndexNode, NodeWithScore, TextNode
from llama_index.core.vector_stores.types import VectorStoreQueryResult
from llama_index.core.node_parser import MarkdownNodeParser

from doc_loader.roche_loader import get_chunk_512_from_content
from node_parser.hierarchical_parser import HierarchicalParser
from src.core.logging import get_logger
from text_splitter.text_splitter2 import RecursiveCharacterTextSplitter2
from text_splitter.zh_text_splitter import ZHSentenceSplitter


def load_document_chunks(file_name: str, json_str: str, doc_id: str) -> pd.DataFrame:
    """加载文档并切分成512大小的块"""
    return pd.concat(
        [
            get_chunk_512_from_content(
                file_name, file_path=json_str, input_type="json_str", doc_id=doc_id
            )
        ],
        ignore_index=True,
    )


def create_big_nodes(chunk_df: pd.DataFrame) -> List[TextNode]:
    """从DataFrame创建大块TextNode列表"""
    return [
        TextNode(
            text=row["all_text"],
            id_=row["id_"],
            metadata={
                # "doc_id": row["md5_name"],
                # "doc_name": row["file_name"],
                "md5_name": row["md5_name"],
                "file_name": row["file_name"],
                "header": row["header"],
                "heading": row["heading"],
                "page_label": row["start_page"],
                "start_page": row["start_page"],
                "end_page": row["end_page"],
            },
        )
        for _, row in chunk_df.iterrows()
    ]


def create_small_nodes(
    big_nodes: List[TextNode], tokenizer: Callable, split_fn: Callable = None
) -> List[IndexNode]:
    """将大块进一步切分成小块"""
    small_nodes = []
#     small_parser = ZHSentenceSplitter(
#         chunk_size=128,
#         chunk_overlap=0,
#         include_prev_next_rel=True,
#         include_metadata=True,
#         tokenizer=tokenizer,
# )
#     split_fn = small_parser.get_nodes_from_documents

    small_parser = RecursiveCharacterTextSplitter2()
    split_fn = small_parser.get_nodes_from_documents

    for big_chunk in big_nodes:
        small_chunks = split_fn([big_chunk])
        for idx, small_chunk in enumerate(small_chunks):
            small_chunk.id_ = f"{big_chunk.node_id}_chunk128_{idx}"
            small_chunk.text = format_chunk_text(small_chunk, big_chunk.metadata)
            small_chunk.metadata["index_text"] = (
                big_chunk.text
            )  # metadata补充写入index_text字段！！！
            # small_chunk.excluded_embed_metadata_keys.extend(list(small_chunk.metadata.keys()))  # 理论上所有metadata都应被排除
            small_chunk.excluded_embed_metadata_keys.append("index_text")
        small_nodes.extend(
            [IndexNode.from_text_node(node, big_chunk.node_id) for node in small_chunks]
        )
    return small_nodes


def format_chunk_text(chunk: TextNode, metadata: Dict) -> str:
    """格式化chunk文本，添加header和heading信息"""
    header = (
        f"header: {metadata['header']}\n\n" if pd.notna(metadata.get("header")) else ""
    )
    heading = (
        f"heading: {metadata['heading']}\n\n"
        if pd.notna(metadata.get("heading"))
        else ""
    )
    return f"{header}{heading}{chunk.text}"


def create_nodes_from_pdf_blocks(
    blocks: List[Dict[str, Any]],
    doc_name: str,
    doc_id: str,
) -> List[IndexNode]:
    """Build block-level nodes from parsed PDF layout blocks."""
    nodes: List[IndexNode] = []
    for idx, block in enumerate(blocks):
        content_type = str(block.get("type", "text")).lower().strip() or "text"
        page_label = block.get("page_label")
        page_key = page_label if page_label is not None else "unknown"
        index_id = f"doc_{doc_id}_page_{page_key}_chunk512_0"
        block_id = block.get("id") or f"doc_{doc_id}_page_{page_key}_chunk128_{idx}"
        block_text = str(block.get("text") or "")
        index_text = str(block.get("index_text") or block_text)

        metadata: Dict[str, Any] = {
            "md5_name": doc_id,
            "file_name": doc_name,
            "index_text": index_text,
            "content_type": content_type,
        }
        if page_label is not None:
            metadata["page_label"] = page_label

        for key in [
            "image_path",
            "image_name",
            "image_width",
            "image_height",
            "bbox",
            "header",
            "heading",
            "source",
        ]:
            value = block.get(key)
            if value is not None:
                metadata[key] = value

        node = IndexNode(
            id_=str(block_id),
            text=block_text,
            index_id=index_id,
            metadata=metadata,
        )
        node.excluded_embed_metadata_keys.append("index_text")
        nodes.append(node)
    return nodes


def chunk_into_small_nodes(
    input_type: str, text: str, doc_name: str, doc_id: str, tokenizer: Callable
) -> List[IndexNode]:
    """将文本切分成小块，其中小块的metadata中包含index_text字段（index_text字段为父节点文本）"""
    logger = get_logger(__name__)

    logger.info(f"Starting document chunking: {doc_name}, input type: {input_type}")

    if input_type == "autopdf":
        # 1. 加载文档并切分段落
        logger.info("Loading document using autopdf format")
        big_chunks_df = load_document_chunks(
            file_name=doc_name, json_str=text, doc_id=doc_id
        )
        logger.info(f"Generated large chunks count: {len(big_chunks_df)}")

        # 2. 创建大块
        big_nodes = create_big_nodes(big_chunks_df)
        logger.info(f"Created large node count: {len(big_nodes)}")

        # 3. 切分小段落
        small_nodes = create_small_nodes(big_nodes, tokenizer)
        logger.info(f"Split small node count: {len(small_nodes)}")

    elif input_type == "raw":
        # 1. 切分大小段落
        logger.info("Loading document using raw format")
        document = Document(text=text)
        document.metadata = {"md5_name": doc_id, "file_name": doc_name}
        document.id_ = doc_id

        all_nodes = HierarchicalParser(
            chunk_sizes=[128, 512],
            tokenizer=tokenizer,
            embedding_model=None,
        ).parse_nodes([document])
        logger.info(f"Hierarchical parsing generated node count: {len(all_nodes)}")

        # 2. 合并大段落到小段落中
        id_to_node = {node.id_: node for node in all_nodes}
        small_nodes = []
        for node in all_nodes:
            if "_chunk128_" in node.id_:
                node.metadata["index_text"] = id_to_node[node.index_id].text
                node.excluded_embed_metadata_keys.append("index_text")
                small_nodes.append(node)
        logger.info(f"Extracted small node count: {len(small_nodes)}")

    elif input_type == "json_list":
        # 1. 切分大小段落
        logger.info("Loading document using json_list format")
        json_list = ast.literal_eval(text)
        small_nodes = []
        for idx, json_dict in enumerate(json_list):
            small_nodes.append(
                IndexNode(
                    id_=f"doc_{doc_id}_chunk512_0_chunk128_{idx}",
                    text=json.dumps(json_dict, ensure_ascii=False),
                    index_id=f"doc_{doc_id}_chunk512_0_chunk128_{idx}",
                    metadata={
                        "md5_name": doc_id,
                        "file_name": doc_name,
                        "index_text": json.dumps(json_dict, ensure_ascii=False)
                    },
                    excluded_embed_metadata_keys=["index_text"]
                )
            )
        logger.info(f"Extracted small node count: {len(small_nodes)}")

    elif input_type == "markdown":
        logger.info("Loading document using markdown format")
        document = Document(text=text)
        document.metadata = {"md5_name": doc_id, "file_name": doc_name}
        document.id_ = doc_id

        # 使用MarkdownNodeParser将文档切分成段落
        # https://docs.llamaindex.ai/en/stable/api_reference/node_parsers/markdown/
        parser = MarkdownNodeParser()
        nodes = parser.get_nodes_from_documents([document])
        for idx, node in enumerate(nodes):
            node.id_ = f"doc_{doc_id}_chunk512_{idx}"

        logger.info(f"Generated large chunks count: {len(nodes)}")

        # 进一步切分段落
        small_nodes = create_small_nodes(nodes, tokenizer)
        logger.info(f"Split small node count: {len(small_nodes)}")

    elif input_type == "pdf_blocks":
        logger.info("Loading document using pdf_blocks format")
        blocks = json.loads(text)
        if not isinstance(blocks, list):
            raise ValueError("pdf_blocks 输入格式错误，期望 JSON list。")
        small_nodes = create_nodes_from_pdf_blocks(
            blocks=blocks,
            doc_name=doc_name,
            doc_id=doc_id,
        )
        logger.info(f"Generated block node count: {len(small_nodes)}")

    if not small_nodes:
        raise ValueError(f"文档切分结果为空: doc_name={doc_name}, input_type={input_type}")
    logger.info(f"Example small node: {small_nodes[0]}")

    return small_nodes


def process_search_results(results: VectorStoreQueryResult) -> List[NodeWithScore]:
    """处理搜索结果：根据index_id进行去重并计算得到父节点"""
    logger = get_logger(__name__)

    logger.info(f"Processing search results, original result count: {len(results.nodes)}")

    index_id_to_node = {}
    for node, score in zip(results.nodes, results.similarities):
        if node.index_id in index_id_to_node:
            continue

        new_node = node.copy()
        new_node.id_ = new_node.index_id
        new_node.text = new_node.metadata.get("index_text", "")
        index_id_to_node[new_node.index_id] = NodeWithScore(node=new_node, score=score)

    logger.info(f"Deduplicated result count: {len(index_id_to_node)}")
    return list(index_id_to_node.values())


def read_excel_to_dict_list(file_path, sheet_name=0):
    """
    读取Excel文件并转换为字典列表，自动处理合并单元格和异常行

    参数:
        file_path: Excel文件路径
        sheet_name: 工作表名称或索引

    返回:
        包含字典的列表，每个字典以表头为键
    """
    # 使用openpyxl处理合并单元格
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if isinstance(sheet_name, int):
        sheet = wb.worksheets[sheet_name]
    else:
        sheet = wb[sheet_name]

    # 处理合并单元格
    merged_cells = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.min_col, merged_range.min_row, merged_range.max_col, merged_range.max_row
        top_left_cell_value = sheet.cell(min_row, min_col).value

        # 将合并单元格的值存储起来
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if row != min_row or col != min_col:
                    merged_cells[(row, col)] = top_left_cell_value

    # 使用pandas读取数据
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

    # 自动检测表头行和需要跳过的行
    header_row, skip_rows = detect_header_and_skip_rows(df)

    if header_row is not None:
        # 设置表头
        headers = []
        for j in range(len(df.columns)):
            # 检查是否为合并单元格
            excel_row = header_row + 1  # pandas索引从0开始，Excel从1开始
            excel_col = j + 1  # pandas列索引从0开始，Excel从1开始

            if (excel_row, excel_col) in merged_cells:
                header_value = merged_cells[(excel_row, excel_col)]
            else:
                header_value = df.iloc[header_row, j]

            headers.append(header_value if pd.notna(header_value) else f"Column_{j}")

        # 过滤掉空表头和无效表头
        valid_headers = []
        valid_indices = []
        for i, header in enumerate(headers):
            if pd.notna(header) and not (isinstance(header, str) and header.startswith('Unnamed:')):
                valid_headers.append(header)
                valid_indices.append(i)

        # 创建结果列表
        result = []

        # 遍历数据行
        for i in range(len(df)):
            # 跳过表头行和需要跳过的行
            if i == header_row or i in skip_rows:
                continue

            row = df.iloc[i]

            # 创建行字典
            row_dict = {}
            for j, header in zip(valid_indices, valid_headers):
                # 检查是否为合并单元格
                excel_row = i + 1  # pandas索引从0开始，Excel从1开始
                excel_col = j + 1  # pandas列索引从0开始，Excel从1开始

                if (excel_row, excel_col) in merged_cells:
                    cell_value = merged_cells[(excel_row, excel_col)]
                else:
                    cell_value = row.iloc[j]

                # 只有当值不为空时才添加到字典中
                if pd.notna(cell_value):
                    row_dict[header] = cell_value

            # 只有当字典不为空且包含足够的键值对时才添加到结果列表中
            if row_dict and len(row_dict) >= len(valid_headers) * 0.3:  # 至少包含30%的有效列
                result.append(row_dict)

        return result

    return []

def detect_header_and_skip_rows(df):
    """
    自动检测表头行和需要跳过的行

    参数:
        df: pandas DataFrame

    返回:
        header_row: 表头行索引
        skip_rows: 需要跳过的行索引列表
    """
    skip_rows = []
    header_row = None

    # 计算每行的非空单元格数量
    row_non_empty_counts = []
    for i in range(len(df)):
        non_empty_count = df.iloc[i].notna().sum()
        row_non_empty_counts.append(non_empty_count)

    # 计算非空单元格数量的统计信息
    non_empty_counts = np.array(row_non_empty_counts)
    median_count = np.median(non_empty_counts)

    # 找出可能的表头行（通常是在数据行之前，且非空单元格数量较多的行）
    potential_header_rows = []
    for i in range(len(df)):
        if row_non_empty_counts[i] >= median_count:
            potential_header_rows.append(i)
            if len(potential_header_rows) > 0 and i > potential_header_rows[0] + 5:
                break

    if potential_header_rows:
        # 选择第一个可能的表头行
        header_row = potential_header_rows[0]

        # 标记表头行之前的所有行为需要跳过的行
        for i in range(header_row):
            skip_rows.append(i)

        # 检测表头行之后的异常行（如分类标题行）
        for i in range(header_row + 1, len(df)):
            # 如果行的非空单元格数量明显少于中位数，可能是标题行或分隔行
            if row_non_empty_counts[i] < median_count * 0.5:
                skip_rows.append(i)
            # 如果行中大部分单元格为空，也可能是标题行
            elif row_non_empty_counts[i] < len(df.columns) * 0.3:
                skip_rows.append(i)

    return header_row, skip_rows


if __name__ == "__main__":
    from doc_loader.roche_loader import get_md5

    file_name = "markdown_demo"
    markdown_text = "# 示例文档\n\n这是一个用于本地调试 chunk_into_small_nodes 的 markdown 示例。"
    doc_id = get_md5(file_name)
    small_nodes = chunk_into_small_nodes(
        input_type="markdown",
        text=markdown_text,
        doc_name=file_name,
        doc_id=doc_id,
        tokenizer=None,
    )
    print(small_nodes)
